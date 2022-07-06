import FreeMobileAPI
import openpyxl
import time
import colorama


EXCEL_SHEET = "/volume1/downloads/FreeMobile/freemobile.xlsx"

doc = openpyxl.load_workbook(EXCEL_SHEET, read_only=False, data_only=False)

closed = False
while closed == False:
    try:
        doc.save(EXCEL_SHEET)
        closed = True         
    except Exception as e:
        print(colorama.Fore.RED, end='')
        print("Excel file was not closed !")
        print("Please close the file before continuing...")

        time.sleep(30)


ws = {}
for worksheet in range(0, len(doc.worksheets)):
    ws[doc.worksheets[worksheet].title] = { "workSheet":doc.worksheets[worksheet], "workSheetIndex":worksheet, "columns": {} }
    for row in doc.worksheets[worksheet].rows:
        for column in row:
            try:
                ws[doc.worksheets[worksheet].title]["columns"][column.value] = column.column
            except:
                pass
        break


for userRow in range(2, ws["User"]["workSheet"].max_row + 1):    
    API = FreeMobileAPI.FreeMobileAPI()
    API.Login(ws["User"]["workSheet"].cell(row=userRow, column=ws["User"]["columns"]["UserId"]).value, ws["User"]["workSheet"].cell(row=userRow, column=ws["User"]["columns"]["Password"]).value)

    for directionRow in range(2, ws["Direction"]["workSheet"].max_row + 1):
        if API.EnableFilter(Enabled=bool(ws["User"]["workSheet"].cell(row=userRow, column=ws["User"]["columns"]["Deny {}".format(ws["Direction"]["workSheet"].cell(row=directionRow, column=ws["Direction"]["columns"]["Direction"]).value)]).value), Direction=ws["Direction"]["workSheet"].cell(row=directionRow, column=ws["Direction"]["columns"]["Value"]).value):
            print(colorama.Fore.LIGHTCYAN_EX, end='')
            print("Set Default Filter {direction} to {value} for {user}".format(direction=ws["Direction"]["workSheet"].cell(row=directionRow, column=ws["Direction"]["columns"]["Direction"]).value, value=bool(ws["User"]["workSheet"].cell(row=userRow, column=ws["User"]["columns"]["Deny {}".format(ws["Direction"]["workSheet"].cell(row=directionRow, column=ws["Direction"]["columns"]["Direction"]).value)]).value), user=ws["User"]["workSheet"].cell(row=userRow, column=ws["User"]["columns"]["User"]).value))


    for filterRow in range(ws["Filter"]["workSheet"].max_row, 1, -1):    
        if ws["Filter"]["workSheet"].cell(row=filterRow, column=ws["Filter"]["columns"]["RuleId"]).value in (None, "") and \
            ws["Filter"]["workSheet"].cell(row=filterRow, column=ws["Filter"]["columns"]["User"]).value == ws["User"]["workSheet"].cell(row=userRow, column=ws["User"]["columns"]["User"]).value:

            params = "rule-id="            

            time = {"start":None, "end":None}
            if ws["Filter"]["workSheet"].cell(row=filterRow, column=ws["Filter"]["columns"]["Time"]).value != None:
                for timeRow in range(2, ws["Time"]["workSheet"].max_row + 1):
                    if ws["Filter"]["workSheet"].cell(row=filterRow, column=ws["Filter"]["columns"]["Time"]).value == ws["Time"]["workSheet"].cell(row=timeRow, column=ws["Time"]["columns"]["Time"]).value:
                        time["start"] = ws["Time"]["workSheet"].cell(row=timeRow, column=ws["Time"]["columns"]["From"]).value
                        time["end"] = ws["Time"]["workSheet"].cell(row=timeRow, column=ws["Time"]["columns"]["To"]).value
                        params += "&start_h={sh}&start_m={sm}&end_h={eh}&end_m={em}".format(sh=time["start"].hour, sm=time["start"].minute, eh=time["end"].hour, em=time["end"].minute)
                        break
            else:
                params += "&full-day=1"


            if ws["Filter"]["workSheet"].cell(row=filterRow, column=ws["Filter"]["columns"]["Day"]).value != None:
                for dayRow in range(2, ws["Time"]["workSheet"].max_row + 1):
                    if ws["Filter"]["workSheet"].cell(row=filterRow, column=ws["Filter"]["columns"]["Day"]).value == ws["Day"]["workSheet"].cell(row=dayRow, column=ws["Day"]["columns"]["Day"]).value:

                        if bool(ws["Day"]["workSheet"].cell(row=dayRow, column=ws["Day"]["columns"]["Sunday"]).value) == True:
                            params += "&days[]=1"
                    
                        if bool(ws["Day"]["workSheet"].cell(row=dayRow, column=ws["Day"]["columns"]["Monday"]).value) == True:
                            params += "&days[]=2"

                        if bool(ws["Day"]["workSheet"].cell(row=dayRow, column=ws["Day"]["columns"]["Tuesday"]).value) == True:
                            params += "&days[]=3"

                        if bool(ws["Day"]["workSheet"].cell(row=dayRow, column=ws["Day"]["columns"]["Wednesday"]).value) == True:
                            params += "&days[]=4"

                        if bool(ws["Day"]["workSheet"].cell(row=dayRow, column=ws["Day"]["columns"]["Thursday"]).value) == True:
                            params += "&days[]=5"

                        if bool(ws["Day"]["workSheet"].cell(row=dayRow, column=ws["Day"]["columns"]["Friday"]).value) == True:
                            params += "&days[]=6"

                        if bool(ws["Day"]["workSheet"].cell(row=dayRow, column=ws["Day"]["columns"]["Saturday"]).value) == True:
                            params += "&days[]=7"

                        break
            else:
                params += "&full-week=1"

            if ws["Filter"]["workSheet"].cell(row=filterRow, column=ws["Filter"]["columns"]["Target"]).value != None:
                for targetRow in range(2, ws["Target"]["workSheet"].max_row + 1):
                    if ws["Filter"]["workSheet"].cell(row=filterRow, column=ws["Filter"]["columns"]["Target"]).value == ws["Target"]["workSheet"].cell(row=targetRow, column=ws["Target"]["columns"]["Target"]).value:
                        params += "&media={value}".format(value=ws["Target"]["workSheet"].cell(row=targetRow, column=ws["Target"]["columns"]["Value"]).value)
                        break


            if ws["Filter"]["workSheet"].cell(row=filterRow, column=ws["Filter"]["columns"]["Action"]).value != None:
                for actionRow in range(2, ws["Action"]["workSheet"].max_row + 1):
                    if ws["Filter"]["workSheet"].cell(row=filterRow, column=ws["Filter"]["columns"]["Action"]).value == ws["Action"]["workSheet"].cell(row=actionRow, column=ws["Action"]["columns"]["Action"]).value:
                        params += "&action={action}".format(action=ws["Action"]["workSheet"].cell(row=actionRow, column=ws["Action"]["columns"]["Value"]).value)
                        break


            if ws["Filter"]["workSheet"].cell(row=filterRow, column=ws["Filter"]["columns"]["Phone"]).value != None:
                params += "&pattern={phone}".format(phone=ws["Filter"]["workSheet"].cell(row=filterRow, column=ws["Filter"]["columns"]["Phone"]).value)

            if ws["Filter"]["workSheet"].cell(row=filterRow, column=ws["Filter"]["columns"]["Description"]).value != None:
                params += "&description={desc}".format(desc=ws["Filter"]["workSheet"].cell(row=filterRow, column=ws["Filter"]["columns"]["Description"]).value)


            if ws["Filter"]["workSheet"].cell(row=filterRow, column=ws["Filter"]["columns"]["Direction"]).value != None:
                for directionRow in range(2, ws["Direction"]["workSheet"].max_row + 1):
                    if ws["Filter"]["workSheet"].cell(row=filterRow, column=ws["Filter"]["columns"]["Direction"]).value == ws["Direction"]["workSheet"].cell(row=directionRow, column=ws["Direction"]["columns"]["Direction"]).value:
                        params += "&direction={}".format(ws["Direction"]["workSheet"].cell(row=directionRow, column=ws["Direction"]["columns"]["Value"]).value)
                        break


            if API.AddFilter(params):
                print(colorama.Fore.LIGHTYELLOW_EX, end='')
                print("{action} {target} {direction} for {user} to #{phone}".format(direction=ws["Filter"]["workSheet"].cell(row=filterRow, column=ws["Filter"]["columns"]["Direction"]).value, action=ws["Filter"]["workSheet"].cell(row=filterRow, column=ws["Filter"]["columns"]["Action"]).value, target=ws["Filter"]["workSheet"].cell(row=filterRow, column=ws["Filter"]["columns"]["Target"]).value, user=ws["Filter"]["workSheet"].cell(row=filterRow, column=ws["Filter"]["columns"]["User"]).value, phone=ws["Filter"]["workSheet"].cell(row=filterRow, column=ws["Filter"]["columns"]["Phone"]).value))
                ids = API.GetFilterIds()
                ws["Filter"]["workSheet"].cell(row=filterRow, column=ws["Filter"]["columns"]["RuleId"]).value = ids[ws["Filter"]["workSheet"].cell(row=filterRow, column=ws["Filter"]["columns"]["Direction"]).value][len(ids[ws["Filter"]["workSheet"].cell(row=filterRow, column=ws["Filter"]["columns"]["Direction"]).value]) - 1]

            doc.save(EXCEL_SHEET)
            
        elif ws["Filter"]["workSheet"].cell(row=filterRow, column=ws["Filter"]["columns"]["RuleId"]).font.strike == True:
            for userRow in range(2, ws["User"]["workSheet"].max_row + 1):
                if ws["Filter"]["workSheet"].cell(row=filterRow, column=ws["Filter"]["columns"]["User"]).value == ws["User"]["workSheet"].cell(row=userRow, column=ws["User"]["columns"]["User"]).value:
                    API.Login(ws["User"]["workSheet"].cell(row=userRow, column=ws["User"]["columns"]["UserId"]).value, ws["User"]["workSheet"].cell(row=userRow, column=ws["User"]["columns"]["Password"]).value)                                
                    break

            if API.DeleteFilter(ws["Filter"]["workSheet"].cell(row=filterRow, column=ws["Filter"]["columns"]["RuleId"]).value):       
                print(colorama.Fore.LIGHTMAGENTA_EX, end='')
                print("Remove Filter {target} {direction} from {user} for #{phone}".format(target=ws["Filter"]["workSheet"].cell(row=filterRow, column=ws["Filter"]["columns"]["Target"]).value, direction=ws["Filter"]["workSheet"].cell(row=filterRow, column=ws["Filter"]["columns"]["Direction"]).value, user=ws["User"]["workSheet"].cell(row=userRow, column=ws["User"]["columns"]["User"]).value, phone=ws["Filter"]["workSheet"].cell(row=filterRow, column=ws["Filter"]["columns"]["Phone"]).value))
                ws["Filter"]["workSheet"].delete_rows(filterRow, 1)
                doc.save(EXCEL_SHEET)


    API.Logout()


doc.close()

quit(0)
